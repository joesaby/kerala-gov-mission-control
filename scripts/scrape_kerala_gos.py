"""
scrape_kerala_gos.py
--------------------
Scrape Government Orders from the Kerala Government Document Portal
(https://document.kerala.gov.in) issued on or after the date the
V.D. Satheesan-led UDF cabinet was sworn in (18 May 2026) and write
them to an .xlsx + .csv with citations.

Usage:
    pip install requests beautifulsoup4 openpyxl
    python scrape_kerala_gos.py

Output:
    Kerala_GOs_since_18May2026.xlsx
    Kerala_GOs_since_18May2026.csv

Notes:
- The portal's default landing page lists GOs from the last 30 days.
  We paginate by following the "Load more documents" pattern if the
  site exposes one; otherwise a single fetch of the landing page is
  enough for the small post-swearing-in window.
- The portal serves Malayalam + English subject lines. We keep the
  raw text as published.
- Be polite: 1-second delay between requests, identify yourself in
  the User-Agent.
"""

from __future__ import annotations

import csv
import re
import sys
import time
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------- Configuration ----------------------------------------------------

PORTAL_BASE = "https://document.kerala.gov.in"
GO_LIST_URL = f"{PORTAL_BASE}/documentdetails/en/dDVtK21nV2l6c0RxcCtWTC9oTGcvZz09"

# Date on which the new Kerala government took office.
NEW_GOVT_OATH = date(2026, 5, 18)

OUT_DIR = Path("./out")
OUT_DIR.mkdir(exist_ok=True)
XLSX_PATH = OUT_DIR / "Kerala_GOs_since_18May2026.xlsx"
CSV_PATH = OUT_DIR / "Kerala_GOs_since_18May2026.csv"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; KeralaGOResearchBot/1.0; "
        "+contact: research@example.org)"
    ),
    "Accept-Language": "en-IN,en;q=0.9",
}

REQUEST_DELAY_SEC = 1.0


# ---------- Data model -------------------------------------------------------

@dataclass
class GovtOrder:
    sno: int
    date: str            # dd-mm-yyyy as published
    go_number: str       # e.g. "G.O. (RT)300/2026/CLAD"
    department: str
    subject: str
    pdf_url: str

    @property
    def date_obj(self) -> date | None:
        try:
            return datetime.strptime(self.date, "%d-%m-%Y").date()
        except ValueError:
            return None


# ---------- Scraping ---------------------------------------------------------

def fetch(url: str) -> str:
    """Polite GET that raises on HTTP errors."""
    time.sleep(REQUEST_DELAY_SEC)
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return resp.text


# A GO block on the page roughly looks like:
#   subject line ... G.O. (RT)300/2026/CLAD ...
#   <a>Department Name</a>
#   dd-mm-yyyy
#   <a href=".../govtorder...pdf" title="Download">...
GO_NUMBER_RE = re.compile(
    r"G\.O\.\s*\([A-Za-z/&]+\)\s*\d+/\d{4}/[A-Za-z&]+", re.IGNORECASE
)
DATE_RE = re.compile(r"\b(\d{2}-\d{2}-\d{4})\b")


def parse_orders(html: str) -> list[GovtOrder]:
    """Extract GO records from the listing HTML."""
    soup = BeautifulSoup(html, "html.parser")
    orders: list[GovtOrder] = []

    # Every order has a download link to a /documents/governmentorders/*.pdf
    pdf_links = soup.select('a[href*="/documents/governmentorders/"]')

    seen_urls: set[str] = set()
    sno = 0

    for a in pdf_links:
        href = a.get("href", "").strip()
        if not href.endswith(".pdf") or href in seen_urls:
            continue
        seen_urls.add(href)

        # Walk up to the nearest container that holds the subject + dept + date.
        block = a
        for _ in range(6):
            if block.parent is None:
                break
            block = block.parent
            text = block.get_text(" ", strip=True)
            if GO_NUMBER_RE.search(text) and DATE_RE.search(text):
                break

        text = block.get_text(" ", strip=True)

        go_match = GO_NUMBER_RE.search(text)
        date_match = DATE_RE.search(text)
        if not (go_match and date_match):
            continue

        go_number = go_match.group(0).strip()
        date_str = date_match.group(1)

        # Department is usually the first <a> inside the block whose href
        # contains "/deptdocumentdetails/".
        dept_link = block.find(
            "a", href=lambda h: h and "/deptdocumentdetails/" in h
        )
        department = dept_link.get_text(strip=True) if dept_link else ""

        # Subject = text BEFORE the G.O. number on the same block.
        subject = text.split(go_match.group(0))[0]
        subject = re.sub(r"^[\s>•:-]+", "", subject).strip(" .")

        sno += 1
        orders.append(
            GovtOrder(
                sno=sno,
                date=date_str,
                go_number=go_number,
                department=department,
                subject=subject,
                pdf_url=href if href.startswith("http") else PORTAL_BASE + href,
            )
        )

    return orders


def collect_orders_since(cutoff: date) -> list[GovtOrder]:
    print(f"[+] Fetching {GO_LIST_URL}", file=sys.stderr)
    html = fetch(GO_LIST_URL)
    parsed = parse_orders(html)
    print(f"[+] Parsed {len(parsed)} raw orders from landing page",
          file=sys.stderr)

    filtered = [o for o in parsed if o.date_obj and o.date_obj >= cutoff]
    # Re-number after filtering, newest first
    filtered.sort(key=lambda o: (o.date_obj or date.min), reverse=True)
    for i, o in enumerate(filtered, start=1):
        o.sno = i
    print(f"[+] Kept {len(filtered)} orders dated on/after {cutoff.isoformat()}",
          file=sys.stderr)
    return filtered


# ---------- Writers ----------------------------------------------------------

def write_csv(orders: Iterable[GovtOrder], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(
            ["S.No.", "Date", "G.O. Number", "Department",
             "Subject", "Download URL"]
        )
        for o in orders:
            w.writerow(
                [o.sno, o.date, o.go_number, o.department,
                 o.subject, o.pdf_url]
            )
    print(f"[+] Wrote {path}", file=sys.stderr)


def write_xlsx(orders: list[GovtOrder], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GOs since 18-May-2026"

    ws["A1"] = "Government Orders – Government of Kerala"
    ws["A2"] = (f"Period: {NEW_GOVT_OATH.strftime('%d %b %Y')} "
                "(oath of Satheesan-led UDF cabinet) onwards")
    ws["A3"] = f"Source: {GO_LIST_URL}"
    ws["A4"] = f"Compiled: {date.today().strftime('%d %b %Y')}"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    for r in (2, 3, 4):
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10)

    header_row = 6
    headers = ["S.No.", "Date", "G.O. Number", "Department",
               "Subject", "Download URL"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E78")
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, o in enumerate(orders, start=1):
        row = header_row + i
        ws.cell(row=row, column=1, value=o.sno)
        ws.cell(row=row, column=2, value=o.date)
        ws.cell(row=row, column=3, value=o.go_number)
        ws.cell(row=row, column=4, value=o.department)
        ws.cell(row=row, column=5, value=o.subject)
        link = ws.cell(row=row, column=6, value=o.pdf_url)
        link.hyperlink = o.pdf_url
        link.font = Font(name="Arial", color="0563C1",
                         underline="single", size=10)
        for c in range(1, 7):
            cell = ws.cell(row=row, column=c)
            if c != 6:
                cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {1: 6, 2: 13, 3: 26, 4: 22, 5: 70, 6: 55}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = "A7"

    wb.save(path)
    print(f"[+] Wrote {path}", file=sys.stderr)


# ---------- Main -------------------------------------------------------------

def main() -> int:
    orders = collect_orders_since(NEW_GOVT_OATH)
    if not orders:
        print("[!] No orders found in the new-government window.",
              file=sys.stderr)
        return 1
    write_csv(orders, CSV_PATH)
    write_xlsx(orders, XLSX_PATH)
    print(f"[+] Done — {len(orders)} orders.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
